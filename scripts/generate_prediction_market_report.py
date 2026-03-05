from __future__ import annotations

import argparse
import asyncio
import html
import json
import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import config
from services.deepseek import DeepseekClient
from services.twitter_api import TwitterAPIClient

logger = logging.getLogger("prediction_market_report")
LOCAL_TZ = ZoneInfo("Asia/Bangkok")
SEPARATOR = "\n\n---\n\n"
SUMMARY_TYPES = (
    "New Market",
    "Product Update",
    "Partnership",
    "Policy",
    "Regulation",
    "Funding",
    "Event",
    "Other",
)
SUMMARY_TYPE_LOOKUP = {
    "new market": "New Market",
    "new markets": "New Market",
    "market": "New Market",
    "market launch": "New Market",
    "market launched": "New Market",
    "product update": "Product Update",
    "product": "Product Update",
    "feature": "Product Update",
    "partnership": "Partnership",
    "partner": "Partnership",
    "policy": "Policy",
    "regulation": "Regulation",
    "regulatory": "Regulation",
    "funding": "Funding",
    "fundraise": "Funding",
    "event": "Event",
    "other": "Other",
}
SUMMARY_PREFIX_RE = re.compile(r"^\[(?P<label>[^\]]+)\]\s*(?P<body>.+)$")
SUMMARY_LABEL_RE = re.compile(
    r"^(?P<label>new market|new markets|market launch|market launched|market|product update|product|feature|partnership|partner|policy|regulation|regulatory|funding|fundraise|event|other)\s*[:\-]\s*(?P<body>.+)$",
    re.IGNORECASE,
)
LEADING_BULLETS_RE = re.compile(r"^(?:[-*•]\s*)+")
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


@dataclass(frozen=True)
class AccountSpec:
    label: str
    username: str


ACCOUNTS: list[AccountSpec] = [
    AccountSpec("Polymarket (@Polymarket)", "Polymarket"),
    AccountSpec("kalshi (@kalshi)", "kalshi"),
    AccountSpec("BuzzingApp (@BuzzingApp)", "BuzzingApp"),
    AccountSpec("MyriadMarkets (@MyriadMarkets)", "MyriadMarkets"),
    AccountSpec("trylimitless (@trylimitless)", "trylimitless"),
    AccountSpec("meleemarkets (@meleemarkets)", "meleemarkets"),
    AccountSpec("42space (@42space)", "42space"),
    AccountSpec("TheHedgehog_io (@TheHedgehog_io)", "TheHedgehog_io"),
    AccountSpec("0xProbable (@0xProbable)", "0xProbable"),
    AccountSpec("opinionlabsxyz (@opinionlabsxyz)", "opinionlabsxyz"),
    AccountSpec("predictdotfun (@predictdotfun)", "predictdotfun"),
    AccountSpec("sportfun (@sportfun)", "sportfun"),
    AccountSpec("Novig (@Novig)", "Novig"),
    AccountSpec("Seer(@seer_pm)", "seer_pm"),
    AccountSpec("Kairos (KairosTradeX)", "KairosTradeX"),
    AccountSpec("The Clearing Company", "theclearingco"),
    AccountSpec("XO Market(@xodotmarket)", "xodotmarket"),
]


def parse_args() -> argparse.Namespace:
    default_output = Path.home() / "Desktop" / "Prediction_Markets_Daily_Track_2026-03-02_to_2026-03-04.xlsx"
    parser = argparse.ArgumentParser(
        description="Generate a daily prediction market Twitter report workbook."
    )
    parser.add_argument("--start-date", default="2026-03-02", help="Inclusive local start date (YYYY-MM-DD).")
    parser.add_argument("--end-date", default="2026-03-04", help="Inclusive local end date (YYYY-MM-DD).")
    parser.add_argument(
        "--output-path",
        default=str(default_output),
        help="Destination .xlsx path.",
    )
    parser.add_argument(
        "--source-workbook",
        default="",
        help="Existing workbook path to reuse the 'Raw Tweets' sheet instead of fetching from Twitter.",
    )
    parser.add_argument("--twitter-token", default=config.TWITTER_TOKEN, help="Twitter API bearer token.")
    parser.add_argument("--twitter-base-url", default=config.TWITTER_API_BASE, help="Twitter API base URL.")
    parser.add_argument("--deepseek-api-key", default=config.DEEPSEEK_API_KEY, help="DeepSeek API key.")
    parser.add_argument("--deepseek-base-url", default=config.DEEPSEEK_API_BASE, help="DeepSeek API base URL.")
    parser.add_argument(
        "--skip-summaries",
        action="store_true",
        help="Skip DeepSeek summary generation and leave summary sheets blank.",
    )
    parser.add_argument(
        "--twitter-concurrency",
        type=int,
        default=4,
        help="Concurrent Twitter account fetches.",
    )
    parser.add_argument(
        "--deepseek-concurrency",
        type=int,
        default=2,
        help="Concurrent DeepSeek summary requests.",
    )
    return parser.parse_args()


def iter_dates(start_day: date, end_day: date) -> list[date]:
    days: list[date] = []
    current = start_day
    while current <= end_day:
        days.append(current)
        current += timedelta(days=1)
    return days


def get_range_bounds(start_day: date, end_day: date) -> tuple[datetime, datetime]:
    now_local = datetime.now(LOCAL_TZ)
    start_dt = datetime.combine(start_day, time.min, tzinfo=LOCAL_TZ)
    if end_day == now_local.date():
        end_dt = now_local
    else:
        end_dt = datetime.combine(end_day, time.max, tzinfo=LOCAL_TZ)
    return start_dt, end_dt


def parse_tweet_datetime(tweet: dict[str, Any]) -> datetime | None:
    for key in ("createdAt", "created_at", "timestamp", "ts"):
        raw = tweet.get(key)
        if not raw:
            continue

        if isinstance(raw, (int, float)):
            value = float(raw)
            if value > 10_000_000_000:
                value /= 1000
            return datetime.fromtimestamp(value, tz=timezone.utc).astimezone(LOCAL_TZ)

        text = str(raw).strip()
        if not text:
            continue

        if text.isdigit():
            value = int(text)
            if value > 10_000_000_000:
                value /= 1000
            return datetime.fromtimestamp(value, tz=timezone.utc).astimezone(LOCAL_TZ)

        iso_candidate = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(iso_candidate)
        except ValueError:
            parsed = None
        if parsed is not None:
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(LOCAL_TZ)

        try:
            parsed = parsedate_to_datetime(text)
        except (TypeError, ValueError, IndexError):
            parsed = None
        if parsed is not None:
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(LOCAL_TZ)

    return None


def normalize_tweet_text(tweet: dict[str, Any]) -> str:
    text = tweet.get("text") or tweet.get("fullText") or tweet.get("full_text") or ""
    text = html.unescape(str(text))
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def get_tweet_sort_key(tweet: dict[str, Any]) -> tuple[datetime, int]:
    created_at = parse_tweet_datetime(tweet) or datetime.fromtimestamp(0, tz=LOCAL_TZ)
    tweet_id = str(tweet.get("id", "0"))
    try:
        numeric_id = int(tweet_id)
    except ValueError:
        numeric_id = 0
    return created_at, numeric_id


def trim_json_block(text: str) -> str:
    match = re.search(r"\{.*\}", text, re.DOTALL)
    return match.group(0) if match else text


def normalize_summary_lines(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        lines: list[str] = []
        for raw_line in value.splitlines():
            line = raw_line.strip()
            if line:
                lines.append(line)
        return lines
    if isinstance(value, list):
        lines: list[str] = []
        for item in value:
            lines.extend(normalize_summary_lines(item))
        return lines
    if isinstance(value, dict):
        lines: list[str] = []
        for key, item in value.items():
            nested_lines = normalize_summary_lines(item)
            if not nested_lines:
                continue
            joined = " ".join(nested_lines).strip()
            if joined:
                lines.append(f"{key}: {joined}")
        return lines
    return [str(value).strip()]


def canonicalize_summary_label(label: str) -> str | None:
    normalized = re.sub(r"[^a-z]+", " ", label.lower()).strip()
    return SUMMARY_TYPE_LOOKUP.get(normalized)


def normalize_summary_line(line: str) -> str:
    text = LEADING_BULLETS_RE.sub("", line.strip())
    if not text:
        return ""

    tagged_match = SUMMARY_PREFIX_RE.match(text)
    if tagged_match:
        label = canonicalize_summary_label(tagged_match.group("label"))
        body = tagged_match.group("body").strip()
        if not body:
            return ""
        return f"[{label or 'Other'}] {body}"

    label_match = SUMMARY_LABEL_RE.match(text)
    if label_match:
        label = canonicalize_summary_label(label_match.group("label"))
        body = label_match.group("body").strip()
        if not body:
            return ""
        return f"[{label or 'Other'}] {body}"

    return f"[Other] {text}"


def normalize_summary_value(value: Any) -> str:
    normalized_lines: list[str] = []
    for line in normalize_summary_lines(value):
        normalized = normalize_summary_line(line)
        if normalized:
            normalized_lines.append(normalized)
    return "\n".join(normalized_lines).strip()


def contains_cjk(text: str) -> bool:
    return bool(CJK_RE.search(text))


async def translate_summary_to_chinese(deepseek: DeepseekClient, summary_text: str) -> str:
    if not summary_text.strip():
        return ""

    system_prompt = (
        "Translate the provided summary lines into Simplified Chinese. "
        "Preserve every leading [Category] tag exactly as-is in English. "
        "Only translate the text after each tag. "
        "Return plain text lines only, one line per summary. "
        "Do not add bullets, numbering, code fences, or explanations."
    )
    translated = await deepseek.chat(
        user_message=summary_text,
        system_prompt=system_prompt,
        temperature=0.0,
        max_tokens=400,
    )
    return normalize_summary_value(translated)


async def fetch_account_tweets(
    client: TwitterAPIClient,
    account: AccountSpec,
    start_dt: datetime,
    end_dt: datetime,
    date_columns: list[date],
) -> dict[date, str]:
    daily: dict[date, list[str]] = {day: [] for day in date_columns}

    results = await asyncio.gather(
        client.get_user_tweets(
            username=account.username,
            max_results=100,
            include_replies=False,
            include_retweets=False,
        ),
        client.search_advanced(
            from_user=account.username,
            max_results=100,
        ),
        return_exceptions=True,
    )

    merged: dict[str, dict[str, Any]] = {}
    fallback_index = 0
    for result in results:
        if isinstance(result, Exception):
            logger.warning("Fetch failed for %s: %s", account.username, result)
            continue

        tweets = result.get("data") or result.get("tweets") or []
        for tweet in tweets:
            text = normalize_tweet_text(tweet)
            created_at = parse_tweet_datetime(tweet)
            if not text or created_at is None:
                continue
            if created_at < start_dt or created_at > end_dt:
                continue
            key = str(tweet.get("id", "")).strip()
            if not key:
                fallback_index += 1
                key = f"no-id-{fallback_index}-{hash((account.username, text, created_at.isoformat()))}"
            existing = merged.get(key)
            if existing is None:
                merged[key] = dict(tweet)
            else:
                merged[key] = {**existing, **tweet}

    for tweet in sorted(merged.values(), key=get_tweet_sort_key):
        created_at = parse_tweet_datetime(tweet)
        if created_at is None:
            continue
        day = created_at.date()
        if day not in daily:
            continue
        daily[day].append(normalize_tweet_text(tweet))

    return {day: SEPARATOR.join(values) if values else "" for day, values in daily.items()}


async def summarize_tweet_block(
    deepseek: DeepseekClient,
    account: str,
    day: date,
    raw_text: str,
) -> tuple[str, str]:
    if not raw_text.strip():
        return "", ""

    system_prompt = (
        "You are producing daily intelligence summaries for prediction market company tweets. "
        "Return strict JSON with keys 'en' and 'zh'. "
        "Each value must be an array of 1-5 short summary strings. "
        "Every summary string must begin with exactly one category tag from this list: "
        f"{', '.join(f'[{item}]' for item in SUMMARY_TYPES)}. "
        "After the tag, add a concise one-line summary. "
        "For 'en', write the summary text in English. "
        "For 'zh', write the summary text in Simplified Chinese. "
        "Use English category tags for both 'en' and 'zh' values. "
        "Do not use bullet markers, numbering, or markdown inside the strings. "
        "focus on product launches, new markets, partnerships, policy, funding, regulation, or other notable announcements. "
        "Do not add markdown code fences. "
        "If the tweets are repetitive, compress repeated points. "
        "If the tweets are low-signal, keep it to 1-3 bullets."
    )
    user_message = (
        f"Account: {account}\n"
        f"Date: {day.isoformat()} ({LOCAL_TZ.key})\n\n"
        f"Tweets:\n{raw_text}"
    )
    response = await deepseek.chat(
        user_message=user_message,
        system_prompt=system_prompt,
        temperature=0.2,
        max_tokens=500,
    )
    parsed = json.loads(trim_json_block(response))
    en = normalize_summary_value(parsed.get("en", ""))
    zh = normalize_summary_value(parsed.get("zh", ""))
    if zh and not contains_cjk(zh):
        zh = await translate_summary_to_chinese(deepseek, zh)
    return en, zh


def style_sheet(ws, date_columns: list[date]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 90
    ws["A1"] = "Prediction Market Handle"
    for idx, day in enumerate(date_columns, start=2):
        ws.cell(row=1, column=idx, value=day.isoformat())

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.column_dimensions["A"].width = 32
    for idx in range(2, len(date_columns) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 65

    for row_idx, account in enumerate(ACCOUNTS, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=account.label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        label_cell.border = border
        for col_idx in range(2, len(date_columns) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row_idx].height = 120

    ws.auto_filter.ref = f"A1:{get_column_letter(len(date_columns) + 1)}{len(ACCOUNTS) + 1}"


def build_workbook(
    raw_rows: dict[str, dict[date, str]],
    en_rows: dict[str, dict[date, str]],
    zh_rows: dict[str, dict[date, str]],
    date_columns: list[date],
) -> Workbook:
    wb = Workbook()
    raw_ws = wb.active
    raw_ws.title = "Raw Tweets"
    en_ws = wb.create_sheet("Twitter Summaries (EN)")
    zh_ws = wb.create_sheet("Twitter Summaries (ZH)")

    for ws in (raw_ws, en_ws, zh_ws):
        style_sheet(ws, date_columns)

    for row_idx, account in enumerate(ACCOUNTS, start=2):
        for col_idx, day in enumerate(date_columns, start=2):
            raw_ws.cell(row=row_idx, column=col_idx, value=raw_rows[account.label][day])
            en_ws.cell(row=row_idx, column=col_idx, value=en_rows[account.label][day])
            zh_ws.cell(row=row_idx, column=col_idx, value=zh_rows[account.label][day])

    return wb


def load_raw_rows_from_workbook(workbook_path: Path) -> tuple[list[date], dict[str, dict[date, str]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        if "Raw Tweets" not in workbook.sheetnames:
            raise ValueError(f"'Raw Tweets' sheet not found in {workbook_path}")

        ws = workbook["Raw Tweets"]
        date_columns: list[date] = []
        for col_idx in range(2, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value in (None, ""):
                continue
            if isinstance(value, datetime):
                date_columns.append(value.date())
                continue
            if isinstance(value, date):
                date_columns.append(value)
                continue
            date_columns.append(date.fromisoformat(str(value).strip()))

        if not date_columns:
            raise ValueError(f"No date columns found in {workbook_path}")

        row_lookup: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if label not in (None, ""):
                row_lookup[str(label).strip()] = row_idx

        raw_rows: dict[str, dict[date, str]] = {
            account.label: {day: "" for day in date_columns}
            for account in ACCOUNTS
        }
        for account in ACCOUNTS:
            row_idx = row_lookup.get(account.label)
            if row_idx is None:
                continue
            for col_idx, day in enumerate(date_columns, start=2):
                value = ws.cell(row=row_idx, column=col_idx).value
                raw_rows[account.label][day] = "" if value is None else str(value).strip()

        return date_columns, raw_rows
    finally:
        workbook.close()


async def run(args: argparse.Namespace) -> Path:
    source_workbook = Path(args.source_workbook).expanduser().resolve() if args.source_workbook else None

    if source_workbook is None and not args.twitter_token:
        raise ValueError("Twitter token is required. Pass --twitter-token or set TWITTER_TOKEN.")
    if not args.skip_summaries and not args.deepseek_api_key:
        raise ValueError("DeepSeek API key is required unless --skip-summaries is used.")

    output_path = Path(args.output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if source_workbook is not None:
        date_columns, raw_rows = load_raw_rows_from_workbook(source_workbook)
        start_day = date_columns[0]
        end_day = date_columns[-1]
        logger.info("Loaded raw tweets from %s", source_workbook)
    else:
        start_day = date.fromisoformat(args.start_date)
        end_day = date.fromisoformat(args.end_date)
        if end_day < start_day:
            raise ValueError("end-date must be on or after start-date.")

        date_columns = iter_dates(start_day, end_day)
        start_dt, end_dt = get_range_bounds(start_day, end_day)
        raw_rows = {}

    twitter_client = None if source_workbook is not None else TwitterAPIClient(args.twitter_base_url, args.twitter_token)
    deepseek_client = (
        None
        if args.skip_summaries
        else DeepseekClient(args.deepseek_api_key, args.deepseek_base_url)
    )

    en_rows: dict[str, dict[date, str]] = {}
    zh_rows: dict[str, dict[date, str]] = {}

    try:
        if twitter_client is not None:
            twitter_semaphore = asyncio.Semaphore(max(1, args.twitter_concurrency))

            async def fetch_one(account: AccountSpec) -> tuple[str, dict[date, str]]:
                async with twitter_semaphore:
                    logger.info("Fetching tweets for %s", account.label)
                    return account.label, await fetch_account_tweets(
                        twitter_client,
                        account,
                        start_dt,
                        end_dt,
                        date_columns,
                    )

            raw_results = await asyncio.gather(*(fetch_one(account) for account in ACCOUNTS))
            raw_rows = {label: rows for label, rows in raw_results}

        if deepseek_client is None:
            for account in ACCOUNTS:
                en_rows[account.label] = {day: "" for day in date_columns}
                zh_rows[account.label] = {day: "" for day in date_columns}
        else:
            deepseek_semaphore = asyncio.Semaphore(max(1, args.deepseek_concurrency))

            async def summarize_one(account: AccountSpec, day: date, raw_text: str) -> tuple[str, date, str, str]:
                if not raw_text.strip():
                    return account.label, day, "", ""
                async with deepseek_semaphore:
                    logger.info("Summarizing %s on %s", account.label, day.isoformat())
                    en_text, zh_text = await summarize_tweet_block(
                        deepseek_client,
                        account.label,
                        day,
                        raw_text,
                    )
                    return account.label, day, en_text, zh_text

            jobs = [
                summarize_one(account, day, raw_rows[account.label][day])
                for account in ACCOUNTS
                for day in date_columns
            ]
            summary_results = await asyncio.gather(*jobs)

            for account in ACCOUNTS:
                en_rows[account.label] = {day: "" for day in date_columns}
                zh_rows[account.label] = {day: "" for day in date_columns}

            for label, day, en_text, zh_text in summary_results:
                en_rows[label][day] = en_text
                zh_rows[label][day] = zh_text

        workbook = build_workbook(raw_rows, en_rows, zh_rows, date_columns)
        workbook.save(output_path)
        logger.info("Workbook saved to %s", output_path)
        return output_path
    finally:
        if twitter_client is not None:
            await twitter_client.close()
        if deepseek_client is not None:
            await deepseek_client.close()


def main() -> None:
    args = parse_args()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    output_path = asyncio.run(run(args))
    print(output_path)


if __name__ == "__main__":
    main()
